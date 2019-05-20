import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active

output = openpyxl.Workbook()
outputsheet = output.active

print('Inverting cells')
for x in range(1, sheet.max_row+1):
    for y in range(1, sheet.max_column + 1):
        outputsheet.cell(row=y, column=x).value = sheet.cell(row=x, column=y).value

output.save('invert_project.xlsx')