import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active

files = [str(e) for e in input().split(' ')]

for file_index, file in enumerate(files):
    with open(file, 'r') as open_file:
        for index, line in enumerate(open_file):
            line = line.strip()
            sheet[get_column_letter(file_index+1) + str(index+1)] = line


wb.save('project.xlsx')