import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

n = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb.active

column_letter = get_column_letter(n)

for i in range(2, n+2):
    column_style = sheet['A' + str(i)]
    column_style.font = Font(bold=True)
    sheet['A' + str(i)] = i-1
    row_style = sheet[str(get_column_letter(i)) + '1']
    row_style.font = Font(bold=True)
    sheet[str(get_column_letter(i)) + '1'] = i-1
    for j in range(2, n+2):
        sheet[str(get_column_letter(i)) + str(j)] = (i-1) * (j-1)

wb.save('mt.xlsx')
