# Setting Row Height and Column Width
# Worksheet objects have row_dimensions and column_dimensions attributes that control row heights and column widths.

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
sheet.row_dimensions[5].height = 70
wb.save('dimensions.xlsx')

# Merging and Unmerging Cells
# A rectangular area of cells can be merged into a single cell with the merge_cells() sheet method.
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')

# To unmerge cells, call the unmerge_cells() sheet method. Enter this into the interactive shell.
wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')