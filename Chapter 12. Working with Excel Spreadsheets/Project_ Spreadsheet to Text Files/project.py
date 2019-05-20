import openpyxl

wb = openpyxl.load_workbook('project.xlsx')
sheet = wb.active


for colNum in range(1, sheet.max_column+1):
    open_file = open('column' + str(colNum) + '.txt', 'w')
    for rowNum in range(1, sheet.max_row+1):
        line = sheet.cell(row=rowNum, column=colNum).value
        if not line:
            continue
        else:
            open_file.write(line + '\n')
    open_file.close()
