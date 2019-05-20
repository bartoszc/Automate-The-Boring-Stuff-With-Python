import sys
import openpyxl
from openpyxl.utils import get_column_letter

n, m, filename = int(sys.argv[1]), int(sys.argv[2]), str(sys.argv[3])

wb = openpyxl.load_workbook(filename)
sheet = wb.active

for rowOfCellObjects in sheet['A' + str(n):str(get_column_letter(sheet.max_column)) + str(m+n-1)]:
    for cellObj in rowOfCellObjects:
        cellObj.value = ''

wb.save('new_myProduce.xlsx')
