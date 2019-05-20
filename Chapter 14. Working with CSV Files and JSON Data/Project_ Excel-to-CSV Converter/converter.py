import os
import openpyxl
import csv

for excelFile in os.listdir('.'):
    if not excelFile.endswith('.xlsx'):
        continue  # skip non-csv files

    wb = openpyxl.load_workbook(excelFile)
    for sheetName in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb[sheetName]

        csvFilename = excelFile[:-5] + '_' + sheetName
        csvFileObj = open(csvFilename, 'w', newline='')
        # Loop through every row in the sheet.
        all_data = []
        for rowNum in range(1, sheet.max_row + 1):
            # Loop through each cell in the row.
            rowData = []
            for colNum in range(1, sheet.max_column + 1):
                rowData.append(sheet.cell(rowNum, colNum).value)

            all_data.append(rowData)

        csvWriter = csv.writer(csvFileObj)
        for row in all_data:
            csvWriter.writerow(row)
        csvFileObj.close()

